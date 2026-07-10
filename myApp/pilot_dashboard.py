"""Helpers for the custom Airamed pilot survey dashboard."""

from __future__ import annotations

import io
from datetime import datetime
from statistics import mean

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .models import PilotSurveyResponse

FIELD_LABELS = {
    'patient_understanding': 'Patient understanding (1–5)',
    'caregiver_confidence': 'Caregiver confidence (1–5)',
    'med_why': 'Understands why taking each medication',
    'med_when': 'Understands when to take medication',
    'med_side_effects': 'Understands possible side effects',
    'med_call_doctor': 'Knows when to call doctor',
    'repeat_call': 'Repeat clarification call',
    'repeat_topic_medication': 'Repeat topic: medication',
    'repeat_topic_wound': 'Repeat topic: wound care',
    'repeat_topic_diet': 'Repeat topic: diet',
    'repeat_topic_appointment': 'Repeat topic: appointments',
    'repeat_topic_activity': 'Repeat topic: activity restrictions',
    'followup_attended': 'Follow-up attended (30 days)',
    'satisfaction_understanding': 'Airamed helped understanding (1–5)',
    'recommend': 'Would recommend Airamed',
    'anxiety': 'Airamed reduced anxiety',
    'staff_education_easier': 'Discharge education easier (1–5)',
    'staff_fewer_questions': 'Fewer repeated questions (1–5)',
    'staff_continue': 'Would continue using Airamed',
    'time_saved': 'Time saved during discharge (minutes)',
    'story': 'Qualitative story',
}

EXPORT_COLUMNS = [
    ('id', 'Response ID'),
    ('created_at', 'Submitted'),
    ('role', 'Role'),
    *[(key, FIELD_LABELS[key]) for key in FIELD_LABELS],
]

UNDERSTANDING_LABELS = {
    '1': 'Not at all',
    '2': 'Slightly',
    '3': 'Somewhat',
    '4': 'Mostly',
    '5': 'Completely',
}

CONFIDENCE_LABELS = {
    '1': 'Not at all confident',
    '2': 'Slightly confident',
    '3': 'Somewhat confident',
    '4': 'Mostly confident',
    '5': 'Completely confident',
}

YES_NO_LABELS = {'yes': 'Yes', 'no': 'No'}
ANXIETY_LABELS = {'yes': 'Yes', 'somewhat': 'Somewhat', 'no': 'No'}
FOLLOWUP_LABELS = {'yes': 'Yes, attended', 'no': 'No, did not attend', 'pending': 'Not yet / pending'}
TIME_SAVED_LABELS = {'0': '0 minutes', '5': '5 minutes', '10': '10 minutes', '15': '15+ minutes'}


def format_field_value(key: str, value) -> str:
    if value is None or value == '':
        return '—'
    if key == 'patient_understanding':
        return f'{value} — {UNDERSTANDING_LABELS.get(str(value), str(value))}'
    if key == 'caregiver_confidence':
        return f'{value} — {CONFIDENCE_LABELS.get(str(value), str(value))}'
    if key in ('recommend', 'staff_continue', 'repeat_call'):
        return YES_NO_LABELS.get(str(value), str(value).title())
    if key == 'anxiety':
        return ANXIETY_LABELS.get(str(value), str(value).title())
    if key == 'followup_attended':
        return FOLLOWUP_LABELS.get(str(value), str(value).title())
    if key == 'time_saved':
        return TIME_SAVED_LABELS.get(str(value), f'{value} minutes')
    if key.startswith('med_') or key.startswith('repeat_topic_'):
        return 'Yes' if str(value).lower() in ('yes', 'true', '1') else str(value)
    if key in ('satisfaction_understanding', 'staff_education_easier', 'staff_fewer_questions'):
        return f'{value} / 5'
    return str(value)


def response_rows(response: PilotSurveyResponse) -> list[dict[str, str]]:
    rows = []
    for key, label in FIELD_LABELS.items():
        if key in response.responses:
            rows.append({
                'key': key,
                'label': label,
                'value': format_field_value(key, response.responses.get(key)),
            })
    return rows


def _avg_numeric(values: list) -> float | None:
    nums = []
    for value in values:
        try:
            nums.append(float(value))
        except (TypeError, ValueError):
            continue
    return round(mean(nums), 2) if nums else None


def _yes_percent(values: list, target: str = 'yes') -> float | None:
    filtered = [v for v in values if v not in (None, '')]
    if not filtered:
        return None
    hits = sum(1 for v in filtered if str(v).lower() == target)
    return round(100 * hits / len(filtered), 1)


def _med_all_yes_percent(responses_list: list[dict]) -> float | None:
    med_keys = ('med_why', 'med_when', 'med_side_effects', 'med_call_doctor')
    eligible = []
    for data in responses_list:
        if any(data.get(k) for k in med_keys):
            eligible.append(all(str(data.get(k, '')).lower() == 'yes' for k in med_keys))
    if not eligible:
        return None
    return round(100 * sum(eligible) / len(eligible), 1)


def compute_stats(queryset) -> dict:
    all_rows = list(queryset)
    responses_list = [row.responses for row in all_rows]

    by_role = {
        PilotSurveyResponse.ROLE_PATIENT: 0,
        PilotSurveyResponse.ROLE_CAREGIVER: 0,
        PilotSurveyResponse.ROLE_NURSE: 0,
    }
    for row in all_rows:
        by_role[row.role] = by_role.get(row.role, 0) + 1

    patient_rows = [r.responses for r in all_rows if r.role == PilotSurveyResponse.ROLE_PATIENT]
    caregiver_rows = [r.responses for r in all_rows if r.role == PilotSurveyResponse.ROLE_CAREGIVER]
    nurse_rows = [r.responses for r in all_rows if r.role == PilotSurveyResponse.ROLE_NURSE]

    return {
        'total': len(all_rows),
        'by_role': by_role,
        'avg_patient_understanding': _avg_numeric([r.get('patient_understanding') for r in patient_rows]),
        'avg_caregiver_confidence': _avg_numeric([r.get('caregiver_confidence') for r in caregiver_rows]),
        'med_all_yes_percent': _med_all_yes_percent(patient_rows),
        'recommend_yes_percent': _yes_percent([r.get('recommend') for r in patient_rows + caregiver_rows]),
        'repeat_call_yes_percent': _yes_percent([r.get('repeat_call') for r in nurse_rows]),
        'followup_attended_percent': _yes_percent([r.get('followup_attended') for r in nurse_rows]),
        'avg_staff_education': _avg_numeric([r.get('staff_education_easier') for r in nurse_rows]),
        'avg_staff_fewer_questions': _avg_numeric([r.get('staff_fewer_questions') for r in nurse_rows]),
        'staff_continue_yes_percent': _yes_percent([r.get('staff_continue') for r in nurse_rows]),
    }


def filter_queryset(queryset, *, role: str | None = None, q: str | None = None):
    if role:
        queryset = queryset.filter(role=role)
    if q:
        queryset = queryset.filter(responses__icontains=q)
    return queryset


def flatten_response(response: PilotSurveyResponse) -> dict:
    row = {
        'id': response.id,
        'created_at': timezone.localtime(response.created_at).strftime('%Y-%m-%d %H:%M'),
        'role': response.get_role_display(),
    }
    for key in FIELD_LABELS:
        row[key] = format_field_value(key, response.responses.get(key))
    return row


def build_excel(queryset) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Pilot Responses'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1B3B7D')

    headers = [label for _, label in EXPORT_COLUMNS]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical='center', wrap_text=True)

    for response in queryset:
        flat = flatten_response(response)
        ws.append([flat.get(key, '—') for key, _ in EXPORT_COLUMNS])

    for col_idx, (key, _) in enumerate(EXPORT_COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        if key == 'story':
            ws.column_dimensions[letter].width = 48
        elif key in ('created_at', 'role'):
            ws.column_dimensions[letter].width = 22
        else:
            ws.column_dimensions[letter].width = 16

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _pdf_styles():
    styles = getSampleStyleSheet()
    return {
        'title': ParagraphStyle(
            'PilotTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1B3B7D'),
            spaceAfter=10,
        ),
        'section': ParagraphStyle(
            'PilotSection',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#249DA8'),
            spaceBefore=12,
            spaceAfter=6,
        ),
        'body': ParagraphStyle(
            'PilotBody',
            parent=styles['BodyText'],
            fontSize=10,
            leading=14,
        ),
        'meta': ParagraphStyle(
            'PilotMeta',
            parent=styles['BodyText'],
            fontSize=9,
            textColor=colors.HexColor('#5E7084'),
        ),
    }


def build_pdf_report(queryset, *, title: str = 'Airamed Pilot Dashboard — Survey Responses') -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=0.65 * inch,
        rightMargin=0.65 * inch,
        topMargin=0.65 * inch,
        bottomMargin=0.65 * inch,
        title=title,
    )
    styles = _pdf_styles()
    story = [
        Paragraph(title, styles['title']),
        Paragraph(
            f'Generated {timezone.localtime(timezone.now()).strftime("%B %d, %Y at %I:%M %p")} · {queryset.count()} response(s)',
            styles['meta'],
        ),
        Spacer(1, 0.2 * inch),
    ]

    stats = compute_stats(queryset)
    summary_data = [
        ['Total responses', str(stats['total'])],
        ['Patients', str(stats['by_role'].get(PilotSurveyResponse.ROLE_PATIENT, 0))],
        ['Caregivers', str(stats['by_role'].get(PilotSurveyResponse.ROLE_CAREGIVER, 0))],
        ['Nurses / staff', str(stats['by_role'].get(PilotSurveyResponse.ROLE_NURSE, 0))],
        ['Avg patient understanding', str(stats['avg_patient_understanding'] or '—')],
        ['Avg caregiver confidence', str(stats['avg_caregiver_confidence'] or '—')],
        ['Medication all-yes rate', f"{stats['med_all_yes_percent']}%" if stats['med_all_yes_percent'] is not None else '—'],
        ['Recommend yes rate', f"{stats['recommend_yes_percent']}%" if stats['recommend_yes_percent'] is not None else '—'],
    ]
    summary_table = Table(summary_data, colWidths=[2.4 * inch, 4.0 * inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8F6F8')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1A2E4A')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D7E2EA')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.extend([Paragraph('Summary', styles['section']), summary_table, Spacer(1, 0.15 * inch)])

    table_header = ['ID', 'Submitted', 'Role', 'Highlights']
    table_rows = [table_header]
    for response in queryset[:200]:
        highlights = []
        for key in (
            'patient_understanding',
            'caregiver_confidence',
            'recommend',
            'staff_continue',
            'story',
        ):
            if key in response.responses:
                label = FIELD_LABELS[key].split('(')[0].strip()
                highlights.append(f'{label}: {format_field_value(key, response.responses[key])}')
        table_rows.append([
            str(response.id),
            timezone.localtime(response.created_at).strftime('%Y-%m-%d %H:%M'),
            response.get_role_display(),
            '\n'.join(highlights[:3]) or '—',
        ])

    responses_table = Table(table_rows, colWidths=[0.55 * inch, 1.15 * inch, 1.2 * inch, 3.5 * inch], repeatRows=1)
    responses_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B3B7D')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D7E2EA')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.extend([Paragraph('All responses', styles['section']), responses_table])

    doc.build(story)
    return buffer.getvalue()


def build_pdf_detail(response: PilotSurveyResponse) -> bytes:
    buffer = io.BytesIO()
    title = f'Pilot Response #{response.id}'
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=0.75 * inch,
        rightMargin=0.75 * inch,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch,
        title=title,
    )
    styles = _pdf_styles()
    story = [
        Paragraph(title, styles['title']),
        Paragraph(
            f'{response.get_role_display()} · {timezone.localtime(response.created_at).strftime("%B %d, %Y at %I:%M %p")}',
            styles['meta'],
        ),
        Spacer(1, 0.2 * inch),
    ]

    data = [['Question', 'Answer']]
    for row in response_rows(response):
        data.append([row['label'], row['value']])

    table = Table(data, colWidths=[2.8 * inch, 3.7 * inch], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1B3B7D')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D7E2EA')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(table)
    doc.build(story)
    return buffer.getvalue()


def export_filename(prefix: str, ext: str) -> str:
    stamp = datetime.now().strftime('%Y%m%d_%H%M')
    return f'{prefix}_{stamp}.{ext}'
